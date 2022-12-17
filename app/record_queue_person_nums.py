# -*- coding: utf-8 -*-
import os 
try: 
    import openpyxl
except:
    print ('can not find xlsx module, ready to install....')
    os.system('pip3 install openpyxl')
    import openpyxl


def clear_detectionInfo_into_xlsx(path, sheet_name):
    """
    将record文件夹下记录监控人数信息的xlsx内容进行清空（从第二行开始）

    :param path: str 型，记录核酸排队人数的xlsx文件所在的路径
    :param sheet_name: str 型，xlsx表格簿中表格的名称，默认是Sheet1，当前采用固定值：test_sheet
    :return: None
    """
    if not os.path.exists(path):
        print ('-----not find xlsx file!!!! pls check its right path!!!-----')
        return
    
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    
    while sheet.max_row > 1:
        sheet.delete_rows(2)

    workbook.save(path)

    print ('clear all current person_info xlsx content over at begin!')
    

def write_detectionInfo_into_xlsx(path, sheet_name, num):
    """
    将监控区域内识别到的人数个数，定期（app/config/params.yml 设置的定时时间）写入record文件夹下的xlsx标

    :param path: str 型，记录核酸排队人数的xlsx文件所在的路径
    :param sheet_name: str 型，xlsx表格簿中表格的名称，默认是Sheet1，当前采用固定值：test_sheet
    :param num: int 型，通过AI模型识别到的监控区域内的人数
    :return: None
    """
    if not os.path.exists(path):
        print ('-----not find xlsx file!!!! pls check its right path!!!-----')
        return

    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    max_rows = sheet.max_row # 获取最大行数，计划在该行下方继续添加数据
    if max_rows < 1:
        print ('-----record xlsx file content is wrong!!!-------')
        return 
    
    value = ['第 {} min'.format(max_rows), num]

    for j in range(2): # 只有2列：时间和当前检测到的人数
        sheet.cell(row=max_rows+1, column=j+1, value=str(value[j]))
    
    workbook.save(path)
    print ('append current person_info into xlsx over!')


if __name__ == '__main__':
    person_nums_lst = [10, 15, 12]
    for num in person_nums_lst:
        write_detectionInfo_into_xlsx('../record/detect_result.xlsx', 'test_sheet', num)